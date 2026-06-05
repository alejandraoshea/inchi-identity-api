import sys
import openpyxl
from pathlib import Path
from typing import Dict, List
from dataclasses import dataclass

@dataclass
class HeadgroupPattern:
    name: str
    smarts: str
    lipid_class: str
    fa_positions: List[str]
    description: str = ""


class PatternGenerator:

    @staticmethod
    def load_templates_from_excel(excel_path: str) -> Dict[str, dict]:
        templates = {}
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["Hoja1"]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 5:
                    continue
                example_num = row[0]
                template_smarts = row[3]
                lipid_type = row[4]
                lipid_subtype = row[5]
                if not template_smarts or not isinstance(template_smarts, str):
                    continue
                if "[G]" not in template_smarts:
                    continue
                template_id = f"template_{example_num}".replace(" ", "_").replace("(", "").replace(")", "")
                templates[template_id] = {
                    "smarts": template_smarts,
                    "lipid_class": lipid_type if lipid_type else "Unknown",
                    "lipid_subtype": lipid_subtype if lipid_subtype else "",
                    "fa_positions": ["N-acyl"] if "SP" in str(lipid_type) else ["sn-1", "sn-2"],
                    "example_num": example_num
                }
            print(f"Loaded {len(templates)} templates from Excel", file=sys.stderr)
        except Exception as e:
            print(f"Error loading templates: {e}", file=sys.stderr)
        return templates

    @staticmethod
    def load_sugars_from_excel(excel_path: str, sheet_name: str = None) -> Dict[str, str]:
        sugars = {}
        try:
            wb = openpyxl.load_workbook(excel_path)
            if sheet_name is None:
                for name in ["Azúcares (G)", "Azúcares", "Sugars", "azucares"]:
                    if name in wb.sheetnames:
                        sheet_name = name
                        break
            if sheet_name not in wb.sheetnames:
                print(f"Warning: Sugar sheet not found. Available: {wb.sheetnames}", file=sys.stderr)
                return {}
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                sugar_name = str(row[0]).strip()
                sugar_smarts = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if sugar_name and sugar_smarts:
                    sugars[sugar_name] = sugar_smarts
            if sugars:
                sugar_list = ', '.join(list(sugars.keys())[:5])
                print(f"Loaded {len(sugars)} sugars: {sugar_list}...", file=sys.stderr)
        except Exception as e:
            print(f"Error reading sugars: {e}", file=sys.stderr)
        return sugars

    @staticmethod
    def generate_patterns(templates: Dict, sugars: Dict[str, str]) -> Dict[str, HeadgroupPattern]:
        """
        Generate all pattern combinations: templates x sugars.

        Substitution strategy:
          [G] (first occurrence) → specific sugar SMARTS
          [G] (remaining)        → [#6]  (generic carbon — O-glycoside links
                                           to further sugars in the chain)
          [R]                    → [#6]  (generic carbon — FA chain attachment
                                           point; marks N-acyl or O-acyl site)

        NOTE: [G] and [R] are NOT valid SMARTS atom queries ([G] has no element
        symbol; [R] means "ring atom"). The original templates used them as
        SMILES-style placeholders. This substitution makes the generated SMARTS
        valid and compilable while preserving the headgroup topology.
        """
        patterns = {}
        for template_id, template_info in templates.items():
            template_smarts = template_info["smarts"]
            if "[G]" not in template_smarts:
                continue
            for sugar_name, sugar_smarts in sugars.items():
                complete_smarts = (
                    template_smarts
                    .replace("[G]", sugar_smarts, 1)   # specific sugar at primary site
                    .replace("[G]", "[#6]")             # generic carbon at remaining sites
                    .replace("[R]", "[#6]")             # generic carbon at FA attachment
                )
                safe_sugar_name = (sugar_name
                    .replace(" ", "_").replace("(", "").replace(")", "").replace("‑", "_"))
                pattern_id = f"{template_id}_{safe_sugar_name}"
                lipid_subtype = template_info.get("lipid_subtype", "")
                pattern_name = (
                    f"{lipid_subtype} ({sugar_name})" if lipid_subtype
                    else f"Example {template_info['example_num']} ({sugar_name})"
                )
                patterns[pattern_id] = HeadgroupPattern(
                    name=pattern_name,
                    smarts=complete_smarts,
                    lipid_class=template_info["lipid_class"],
                    fa_positions=template_info["fa_positions"],
                    description=f"Auto-generated from template {template_info['example_num']} + {sugar_name}"
                )
        return patterns


def build_combined_patterns(manual_patterns: Dict[str, HeadgroupPattern],
                            excel_path: str = None) -> Dict[str, HeadgroupPattern]:
    all_patterns = dict(manual_patterns)
    if excel_path and Path(excel_path).exists():
        try:
            templates = PatternGenerator.load_templates_from_excel(excel_path)
            sugars = PatternGenerator.load_sugars_from_excel(excel_path)
            if templates and sugars:
                generated = PatternGenerator.generate_patterns(templates, sugars)
                added = 0
                for pattern_id, pattern in generated.items():
                    if pattern_id not in all_patterns:
                        all_patterns[pattern_id] = pattern
                        added += 1
                print(f"[OK] Total patterns: {len(all_patterns)} "
                      f"({len(manual_patterns)} manual + {added} generated)", file=sys.stderr)
        except Exception as e:
            print(f"Pattern generation failed: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc()
            print(f"Using manual patterns only ({len(manual_patterns)} total)", file=sys.stderr)
    else:
        if excel_path:
            print(f"Excel not found: {excel_path}", file=sys.stderr)
        print(f"Using manual patterns only ({len(all_patterns)} total)", file=sys.stderr)
    return all_patterns